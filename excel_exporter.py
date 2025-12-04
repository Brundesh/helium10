"""
Excel export module for comprehensive analysis reports.
Creates multi-sheet Excel workbooks with formatted data.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from typing import Dict, Any


def create_excel_export(products_data: Dict[str, Any], comparison_df: pd.DataFrame) -> bytes:
    """
    Create comprehensive Excel export with 4 sheets.

    Sheets:
    1. Rankings - All subcategories sorted by total score
    2. Detailed_Metrics - Full breakdown of all metrics
    3. Keyword_Analysis - Top keywords per subcategory (if Magnet data)
    4. Action_Plan - Only GO/PROCEED recommendations with priority ranking

    Args:
        products_data: Dictionary of product data
        comparison_df: Comparison DataFrame

    Returns:
        Excel file as bytes
    """
    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')

    # === SHEET 1: RANKINGS ===
    rankings_df = comparison_df.copy()

    # Simplify for rankings sheet
    rankings_cols = ['Rank', 'Product', 'Total Score', 'Max Score', 'Score %', 'Grade',
                     'Market Size', 'Top 3 Share %', 'Top Seller Reviews', 'Median Price']

    # Add Magnet columns if available
    if 'Search Volume' in rankings_df.columns and rankings_df['Search Volume'].notna().any():
        rankings_cols.extend(['Search Volume', 'Trend %', 'Products Ranking', 'D/S Ratio (Pg 1-2)', 'Success Rate %'])

    rankings_cols.extend(['Red Flags', 'Yellow Flags', 'Green Signals', 'Action', 'Risk Level'])

    # Filter to available columns
    rankings_cols = [col for col in rankings_cols if col in rankings_df.columns]
    rankings_export = rankings_df[rankings_cols]

    rankings_export.to_excel(writer, sheet_name='Rankings', index=False)

    # === SHEET 2: DETAILED METRICS ===
    detailed_rows = []

    for product_name, data in products_data.items():
        xray = data['xray_metrics']
        viability = data['viability']
        magnet_demand = data.get('magnet_demand_metrics')
        magnet_ds = data.get('magnet_ds_ratio')
        flags = data.get('flags')
        rec = data.get('recommendation')

        row = {
            'Subcategory': product_name,

            # Viability scores
            'Total Score': viability['total_score'],
            'Max Score': viability['max_score'],
            'Score %': viability['score_percentage'],
            'Grade': viability['grade'],

            # X-Ray market metrics
            'Estimated Market': xray['market_size']['estimated_total_market'],
            'Top 10 Revenue': xray['market_size']['top_10_revenue'],
            'Top 3 Share %': xray['market_concentration']['top_3_share_percentage'],
            'Total Products': xray['total_products'],
            'Median Price': xray['median_price'],

            # Top seller
            'Top Seller Brand': xray['top_seller']['brand'],
            'Top Seller Price': xray['top_seller']['price'],
            'Top Seller Revenue': xray['top_seller']['revenue'],
            'Top Seller Reviews': xray['top_seller']['reviews'],

            # Ratings
            'Avg Rating Top 20': xray['rating_analysis']['average_rating_top_20'],
            'Min Rating': xray['rating_analysis']['min_rating'],
            'Max Rating': xray['rating_analysis']['max_rating'],

            # Score breakdown
            'Market Size Score': viability['breakdown']['market_size']['score'],
            'Fragmentation Score': viability['breakdown']['market_fragmentation']['score'],
            'Competition Score': viability['breakdown']['competition']['score'],
            'Satisfaction Score': viability['breakdown']['customer_satisfaction']['score'],
            'Price Score': viability['breakdown']['price_viability']['score'],
        }

        # Add Magnet metrics if available
        if magnet_demand is not None:
            row.update({
                'Seed Keyword': magnet_demand['seed_keyword'],
                'Search Volume': magnet_demand['search_volume'],
                'Trend %': magnet_demand['trend'],
                'Total Listings': magnet_demand['competing_products'],
                'Products Ranking': magnet_ds.get('xray_product_count') if magnet_ds else None,
                'Magnet IQ Score': magnet_demand['magnet_iq_score'],
                'Total Related Keywords': magnet_demand['total_related_keywords'],
                'D/S Ratio (Pg 1-2)': magnet_ds.get('ds_ratio', magnet_ds.get('ratio')) if magnet_ds else None,
                'Success Rate %': magnet_ds.get('success_rate') if magnet_ds else None,
                'Demand Score': magnet_ds['demand_score'] if magnet_ds else None,
                'Supply Score': magnet_ds['supply_score'] if magnet_ds else None,
                'Balance Score': magnet_ds['balance_score'] if magnet_ds else None,
                'DS Verdict': magnet_ds['verdict'] if magnet_ds else None,
            })
        else:
            row.update({
                'Seed Keyword': None,
                'Search Volume': None,
                'Trend %': None,
                'Total Listings': None,
                'Products Ranking': None,
                'Magnet IQ Score': None,
                'Total Related Keywords': None,
                'D/S Ratio (Pg 1-2)': None,
                'Success Rate %': None,
                'Demand Score': None,
                'Supply Score': None,
                'Balance Score': None,
                'DS Verdict': None,
            })

        # Add flags
        if flags:
            row.update({
                'Red Flags Count': len(flags['red_flags']),
                'Yellow Flags Count': len(flags['yellow_flags']),
                'Green Signals Count': len(flags['green_signals']),
            })
        else:
            row.update({
                'Red Flags Count': 0,
                'Yellow Flags Count': 0,
                'Green Signals Count': 0,
            })

        # Add recommendation
        if rec:
            row.update({
                'Action': rec['action'],
                'Risk Level': rec['risk_level'],
                'Reasoning': rec['reasoning'],
            })
        else:
            row.update({
                'Action': None,
                'Risk Level': None,
                'Reasoning': None,
            })

        detailed_rows.append(row)

    detailed_df = pd.DataFrame(detailed_rows)
    detailed_df.to_excel(writer, sheet_name='Detailed_Metrics', index=False)

    # === SHEET 3: KEYWORD ANALYSIS ===
    # Only create if we have Magnet data
    has_magnet = any(data.get('magnet_demand_metrics') is not None for data in products_data.values())

    if has_magnet:
        keyword_rows = []

        for product_name, data in products_data.items():
            magnet_demand = data.get('magnet_demand_metrics')

            if magnet_demand is not None:
                # Add seed keyword
                keyword_rows.append({
                    'Subcategory': product_name,
                    'Keyword Type': 'SEED',
                    'Keyword': magnet_demand['seed_keyword'],
                    'Search Volume': magnet_demand['search_volume'],
                    'Trend %': magnet_demand['trend'],
                    'Competitors': magnet_demand['competing_products'],
                })

                # Add related keywords
                for idx, kw in enumerate(magnet_demand['top_related_keywords'], 1):
                    keyword_rows.append({
                        'Subcategory': product_name,
                        'Keyword Type': f'Related #{idx}',
                        'Keyword': kw['keyword'],
                        'Search Volume': kw['volume'],
                        'Trend %': kw['trend'],
                        'Competitors': kw['competitors'],
                    })

        if keyword_rows:
            keyword_df = pd.DataFrame(keyword_rows)
            keyword_df.to_excel(writer, sheet_name='Keyword_Analysis', index=False)

    # === SHEET 4: ACTION PLAN ===
    # Only products with STRONG_GO or PROCEED
    action_rows = []

    for product_name, data in products_data.items():
        rec = data.get('recommendation')

        if rec and rec['action'] in ['STRONG_GO', 'PROCEED']:
            viability = data['viability']
            flags = data.get('flags')

            action_rows.append({
                'Priority': None,  # Will fill after sorting
                'Subcategory': product_name,
                'Action': rec['action'],
                'Score %': viability['score_percentage'],
                'Grade': viability['grade'],
                'Risk Level': rec['risk_level'],
                'Red Flags': len(flags['red_flags']) if flags else 0,
                'Green Signals': len(flags['green_signals']) if flags else 0,
                'Reasoning': rec['reasoning'],
                'Notes': '',  # Empty for user to fill
            })

    if action_rows:
        action_df = pd.DataFrame(action_rows)
        # Sort by score percentage
        action_df = action_df.sort_values('Score %', ascending=False).reset_index(drop=True)
        action_df['Priority'] = range(1, len(action_df) + 1)

        # Reorder columns
        action_df = action_df[['Priority', 'Subcategory', 'Action', 'Score %', 'Grade',
                               'Risk Level', 'Red Flags', 'Green Signals', 'Reasoning', 'Notes']]

        action_df.to_excel(writer, sheet_name='Action_Plan', index=False)

    # Save and format
    writer.close()

    # Apply formatting
    output.seek(0)
    workbook = writer.book

    # Format each sheet
    _format_rankings_sheet(workbook['Rankings'])
    _format_detailed_sheet(workbook['Detailed_Metrics'])
    if has_magnet and 'Keyword_Analysis' in workbook.sheetnames:
        _format_keyword_sheet(workbook['Keyword_Analysis'])
    if 'Action_Plan' in workbook.sheetnames:
        _format_action_plan_sheet(workbook['Action_Plan'])

    # Save formatted workbook
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue()


def _format_rankings_sheet(sheet):
    """Apply formatting to Rankings sheet."""
    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Freeze top row
    sheet.freeze_panes = 'A2'


def _format_detailed_sheet(sheet):
    """Apply formatting to Detailed_Metrics sheet."""
    _format_rankings_sheet(sheet)  # Same formatting


def _format_keyword_sheet(sheet):
    """Apply formatting to Keyword_Analysis sheet."""
    _format_rankings_sheet(sheet)  # Same formatting


def _format_action_plan_sheet(sheet):
    """Apply formatting to Action_Plan sheet with color coding."""
    # Header formatting
    header_fill = PatternFill(start_color='28a745', end_color='28a745', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Color-code actions
    for row in range(2, sheet.max_row + 1):
        action_cell = sheet[f'C{row}']  # Action column
        if action_cell.value == 'STRONG_GO':
            action_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            action_cell.font = Font(bold=True)
        elif action_cell.value == 'PROCEED':
            action_cell.fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')

    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

    # Freeze top row
    sheet.freeze_panes = 'A2'
